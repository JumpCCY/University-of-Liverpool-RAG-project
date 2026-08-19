import openpyxl
import json

UNIVERSITY_NAMES = ["liverpool", "lancaster", "leeds", "newcastle", "sheffield", "nottingham", "manchester", "york"]


def get_header(sheet) -> list:
    """
    Get the header from the excel sheet and return it as a list
    """
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    return headers

def get_data(sheet) -> list:
    """
    Get the data from the excel sheet and return it as a list of tuples
    """
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

def clean_text(value):
    if isinstance(value, str):
        return ' '.join(value
                .replace('\u00a0', ' ')
                .replace('\u2013', '-')
                .replace('\u2019', "'")
                .replace('\u2014', '-')
                .replace('\n', ' ')
                .split())  # ← splits on any whitespace then rejoins with single space
    return value


for name in UNIVERSITY_NAMES:
    path = f"data/raw/{name}_cs_requirements.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet_obj = workbook.worksheets[0]

    qualifications = [] # list of dictionaries with each qualification
    headers = get_header(sheet_obj) # headers in list
    data = get_data(sheet_obj) # list of tuples with each qualifications ex. data[0] = first row
    data = [tuple(clean_text(cell) for cell in row) for row in data]

    for i in range(sheet_obj.max_row-1):
        info = {}
        for h, d in zip(headers, data[i]):
            info[h] = d
        qualifications.append(info)


    with open(f"data/{name}/json/qualifications.json", "w") as f:
        json.dump(qualifications, f, indent=4, ensure_ascii=False)
